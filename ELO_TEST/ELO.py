# ELO代码实现

"""ELO算法的实现"""
import openpyxl
from scipy.stats import binom


def ELO(Ra, Rb, Sa, K):
    # Ra: 玩家A的ELO值
    # Rb: 玩家B的ELO值
    # Sa: 玩家A的比赛结果，1表示胜利，0表示失败
    # K: K值，一般取值为32
    Ea = 1 / (1 + 10 ** ((Rb - Ra) / 400))
    Eb = 1 / (1 + 10 ** ((Ra - Rb) / 400))
    Ra = Ra + K * (Sa - Ea)
    Rb = Rb + K * ((1 - Sa) - Eb)
    return int(Ra), int(Rb)

# 计算胜率
def ELO_win_rate(Ra, Rb):
    Ea = 1 / (1 + 10 ** ((Rb - Ra) / 400))
    Eb = 1 / (1 + 10 ** ((Ra - Rb) / 400))
    return Ea, Eb


def bo_n_ELO(Ra, Rb, A, B, K):
    if A>0:
        for i in range(A):
            Ra, Rb = ELO(Ra, Rb, 1, K)
    if B>0:
        for i in range(B):
            Ra, Rb = ELO(Ra, Rb, 0, K)
    return Ra, Rb

def ELO_BO5(Ra, Rb):
    Ea, Eb = ELO_win_rate(Ra, Rb)
    win_rate = str(round(Ea * 100, 1))
    print("守擂者胜率为：" + win_rate + "%")
    r_3_0 = binom.pmf(3, 3, Ea)
    r_3_1 = binom.pmf(2, 3, Ea) * Ea
    r_3_2 = binom.pmf(2, 4, Ea) * Ea
    r_2_3 = binom.pmf(2, 4, Ea) * (1 - Ea)
    r_1_3 = binom.pmf(1, 3, Ea) * (1 - Ea)
    r_0_3 = binom.pmf(0, 3, Ea)

    print(f'3:0概率为: {round(r_3_0 * 100, 1)}%')
    print(f'3:1概率为: {round(r_3_1 * 100, 1)}%')
    print(f'3:2概率为: {round(r_3_2 * 100, 1)}%')
    print(f'2:3概率为: {round(r_2_3 * 100, 1)}%')
    print(f'1:3概率为: {round(r_1_3 * 100, 1)}%')
    print(f'0:3概率为: {round(r_0_3 * 100, 1)}%')

    result = {
        r_3_0: '3:0',
        r_3_1: '3:1',
        r_3_2: '3:2',
        r_2_3: '2:3',
        r_1_3: '1:3',
        r_0_3: '0:3'
    }
    print(f'综合胜率为：{round(r_3_0+r_3_1+r_3_2, 4) * 100}%')
    return result.get(max(r_3_0, r_3_1, r_3_2, r_2_3, r_1_3, r_0_3))

def ELO_BO3(Ra, Rb):
    Ea, Eb = ELO_win_rate(Ra, Rb)
    win_rate = str(round(Ea * 100, 1))
    print("守擂者胜率为：" + win_rate + "%")
    r_2_0 = binom.pmf(2, 2, Ea)
    r_2_1 = binom.pmf(1, 2, Ea) * Ea
    r_1_2 = binom.pmf(1, 2, Ea) * (1 - Ea)
    r_0_2 = binom.pmf(0, 2, Ea)

    print(f'2:0概率为: {round(r_2_0 * 100, 1)}%')
    print(f'2:1概率为: {round(r_2_1 * 100, 1)}%')
    print(f'1:2概率为: {round(r_1_2 * 100, 1)}%')
    print(f'0:2概率为: {round(r_0_2 * 100, 1)}%')


    result = {
        r_2_0: '2:0',
        r_2_1: '2:1',
        r_1_2: '1:2',
        r_0_2: '0:2'
    }
    print(f'综合胜率为：{round(r_2_0+r_2_1, 4) * 100}%')
    return result.get(max(r_2_0, r_2_1, r_1_2, r_0_2))

def read_elo():
    workbook = openpyxl.load_workbook('../lpl春季赛常规赛积分.xlsx')
    workbook1 = openpyxl.load_workbook('../ELO积分.xlsx')
    sheet = workbook['Sheet1']
    sheet['A1'] = '队伍A'
    sheet['B1'] = '队伍B'
    sheet['C1'] = 'A胜场'
    sheet['D1'] = 'B胜场'

    sheet1 = workbook1['Sheet1']
    sheet1['A1'] = '队伍'
    sheet1['B1'] = 'ELO积分'

    team = {}
    elo = {}
    team_num = 1
    for row in sheet1.values:
        if row[0] == '队伍':
            continue
        team[row[0]] = team_num
        elo[team_num] = row[1]
        team_num += 1

    workbook.save('../lpl春季赛常规赛积分.xlsx')
    workbook1.save('../ELO积分.xlsx')
    return team, elo

def get_elo(t):
    team, elo = read_elo()
    team_num = team[t]
    return elo.get(team_num)


def save_elo(elo):
    workbook1 = openpyxl.load_workbook('../ELO积分.xlsx')
    sheet1 = workbook1['Sheet1']
    sheet1['A1'] = '队伍'
    sheet1['B1'] = 'ELO积分'
    i = 2
    for key in elo.keys():
        sheet1["B" + str(i)] = elo.get(key)
        i += 1
    workbook1.save('../ELO积分.xlsx')


def save_game(A, B, a, b):
    workbook = openpyxl.load_workbook('../lpl春季赛常规赛积分.xlsx')
    sheet = workbook['Sheet1']
    sheet.append([A, B, a, b])
    workbook.save('../lpl春季赛常规赛积分.xlsx')


